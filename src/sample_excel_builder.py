from __future__ import annotations

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


CONTENT_TYPES_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet3.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet4.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet5.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>
"""

ROOT_RELS_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>
"""

WORKBOOK_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="overview" sheetId="1" r:id="rId1"/>
    <sheet name="sources" sheetId="2" r:id="rId2"/>
    <sheet name="joins" sheetId="3" r:id="rId3"/>
    <sheet name="filters" sheetId="4" r:id="rId4"/>
    <sheet name="target_columns" sheetId="5" r:id="rId5"/>
  </sheets>
</workbook>
"""

WORKBOOK_RELS_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet3.xml"/>
  <Relationship Id="rId4" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet4.xml"/>
  <Relationship Id="rId5" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet5.xml"/>
  <Relationship Id="rId6" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
  <Relationship Id="rId7" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
</Relationships>
"""

STYLES_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="1"><fill><patternFill patternType="none"/></fill></fills>
  <borders count="1"><border/></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>
"""

CORE_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>Mapping Template</dc:title>
  <dc:creator>Codex</dc:creator>
</cp:coreProperties>
"""

APP_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>Codex</Application>
</Properties>
"""


class SampleExcelBuilder:
    """Generate a business requirement mapping workbook."""

    def build(self) -> bytes:
        field_rows = [
            ["序号", "字段英文名", "字段中文名", "业务口径说明", "来源系统", "来源表", "来源字段", "DA试算逻辑", "备注"],
            [1, "", "客户编号", "取客户唯一标识", "客户系统", "b_17000_cust_base_info", "ecif_cust_id", "SELECT ecif_cust_id FROM b_17000_cust_base_info", ""],
            [2, "", "客户名称", "客户中文名称", "客户系统", "b_17000_cust_base_info", "cust_name", "cust_name", ""],
            [3, "", "贷款合同号", "贷款合同编号", "信贷系统", "b_88000_ecif_cont_wide", "cont_no", "cont_no", ""],
            [4, "", "贷款余额", "截至统计日贷款本金余额（不含息）", "信贷系统", "b_88000_ecif_cont_wide", "cont_bal", "SUM(cont_bal)", ""],
            [5, "", "贷款发放日期", "合同发放日期 yyyyMMdd", "信贷系统", "b_88000_ecif_cont_wide", "issue_dt", "issue_dt", ""],
            [6, "", "贷款利率", "合同执行年利率", "信贷系统", "b_88000_ecif_cont_wide", "int_rate", "int_rate", ""],
            [7, "", "币种", "贷款币种，默认CNY", "信贷系统", "b_88000_ecif_cont_wide", "currency", "COALESCE(currency,'CNY')", ""],
            [8, "", "客户类型", "客户分类 A/B/C", "客户系统", "b_17000_cust_base_info", "cust_type", "cust_type WHERE cust_type IN ('A','B','C')", ""],
        ]
        source_rows = [
            ["表名", "中文说明", "关键字段", "分区字段", "更新频率"],
            ["b_17000_cust_base_info", "客户基础信息表", "ecif_cust_id, cust_name, cust_type", "dt", "T+1"],
            ["b_88000_ecif_cont_wide", "合同宽表", "cust_id, cont_no, cont_bal, issue_dt, int_rate, currency", "dt", "T+1"],
        ]
        logic_text = (
            "1. 从客户基础信息表获取客户基本信息（客户编号、名称、类型）\n"
            "2. 关联合同宽表获取贷款信息（合同号、余额、日期、利率、币种）\n"
            "3. 关联键：客户编号 = 合同表的 cust_id\n"
            "4. 客户类型仅含 A/B/C 三类\n"
            "5. 按 dt 分区，每日全量覆盖"
        )
        return self._build_workbook(field_rows, source_rows, logic_text)

    def build_enhanced(self) -> bytes:
        field_rows = [
            ["序号", "字段英文名", "字段中文名", "业务口径说明", "来源系统", "来源表", "来源字段", "DA试算逻辑", "备注"],
            [1, "", "产品编号", "托管产品唯一标识", "托管清算", "dwd_custody_clearing_detail_di", "product_id", "product_id", ""],
            [2, "", "产品名称", "产品展示名称", "产品维表", "dim_custody_product_df", "product_name", "product_name", "需通过 product_id 关联维表获取"],
            [3, "", "清算状态", "清算处理状态", "托管清算", "dwd_custody_clearing_detail_di", "clear_status", "clear_status", "需区分 success、fail、processing"],
            [4, "", "交易金额", "交易侧汇总金额", "托管交易", "dwd_custody_trade_detail_di", "trade_amt", "SUM(trade_amt)", "与清算金额做差异核对"],
            [5, "", "清算金额", "清算侧汇总金额", "托管清算", "dwd_custody_clearing_detail_di", "clear_amt", "SUM(clear_amt)", ""],
            [6, "", "差异金额", "清算金额减交易金额", "托管清算", "dwd_custody_clearing_detail_di", "clear_amt", "SUM(clear_amt - trade_amt)", "需关联交易明细表"],
            [7, "", "清算笔数", "清算流水笔数", "托管清算", "dwd_custody_clearing_detail_di", "clear_id", "COUNT(DISTINCT clear_id)", ""],
            [8, "", "失败笔数", "清算失败流水笔数", "托管清算", "dwd_custody_clearing_detail_di", "clear_status", "SUM(CASE WHEN clear_status = 'fail' THEN 1 ELSE 0 END)", ""],
        ]
        source_rows = [
            ["表名", "中文说明", "关键字段", "分区字段", "更新频率"],
            ["dwd_custody_clearing_detail_di", "托管清算明细事实表", "clear_id, trade_id, product_id, account_id, clear_status, clear_amt, clear_dt", "dt", "T+1"],
            ["dwd_custody_trade_detail_di", "托管交易明细事实表", "trade_id, product_id, trade_amt, trade_dt", "dt", "T+1"],
            ["dim_custody_product_df", "托管产品维表", "product_id, product_name, product_type", "dt", "T+1"],
        ]
        logic_text = (
            "1. 以托管清算明细表为主表，统计指定业务日期的清算结果\n"
            "2. 通过 trade_id 关联托管交易明细表，获取交易金额并计算差异金额\n"
            "3. 通过 product_id 关联产品维表，补充产品名称\n"
            "4. 结果按产品编号、产品名称、清算状态汇总\n"
            "5. 所有事实表均使用 dt='${biz_date}' 过滤"
        )
        return self._build_workbook(field_rows, source_rows, logic_text)

    def _build_workbook(self, field_rows: list[list], source_rows: list[list], logic_text: str) -> bytes:
        workbook = Workbook()
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        field_sheet = workbook.active
        field_sheet.title = "字段需求"
        self._append_rows(field_sheet, field_rows, header_fill, header_font, border)
        self._set_widths(field_sheet, [8, 16, 16, 32, 16, 30, 20, 44, 22])

        source_sheet = workbook.create_sheet("源表结构")
        self._append_rows(source_sheet, source_rows, header_fill, header_font, border)
        self._set_widths(source_sheet, [30, 24, 60, 14, 12])

        logic_sheet = workbook.create_sheet("加工逻辑")
        logic_sheet.append(["整体加工逻辑说明"])
        logic_sheet.append([logic_text])
        logic_sheet["A1"].font = Font(bold=True)
        logic_sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        logic_sheet.column_dimensions["A"].width = 90

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _append_rows(self, sheet, rows, header_fill, header_font, border) -> None:
        for row_index, row in enumerate(rows, start=1):
            sheet.append(row)
            for cell in sheet[row_index]:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if row_index == 1:
                    cell.fill = header_fill
                    cell.font = header_font

    def _set_widths(self, sheet, widths: list[int]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[self._col_name(index)].width = width

    def _shared_strings_xml(self, values: list[str]) -> str:
        body = "".join(f"<si><t>{self._escape(value)}</t></si>" for value in values)
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{len(values)}" uniqueCount="{len(values)}">{body}</sst>'
        )

    def _sheet_xml(self, rows: list[list[str]], string_index: dict[str, int]) -> str:
        row_xml: list[str] = []
        for row_num, row in enumerate(rows, start=1):
            cell_xml: list[str] = []
            for col_num, value in enumerate(row, start=1):
                ref = f"{self._col_name(col_num)}{row_num}"
                idx = string_index[value]
                cell_xml.append(f'<c r="{ref}" t="s"><v>{idx}</v></c>')
            row_xml.append(f'<row r="{row_num}">{"".join(cell_xml)}</row>')
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<sheetData>{"".join(row_xml)}</sheetData></worksheet>'
        )

    def _col_name(self, num: int) -> str:
        result = ""
        while num > 0:
            num, rem = divmod(num - 1, 26)
            result = chr(65 + rem) + result
        return result

    def _escape(self, value: str) -> str:
        return (
            value.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
