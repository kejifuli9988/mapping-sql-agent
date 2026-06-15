from __future__ import annotations

from pathlib import Path
import re

import openpyxl


class StandardizationAdvisor:
    """Recommend standardized target field names from local dictionaries."""

    DEFAULT_DICT_PATH = Path(__file__).resolve().parents[1] / "mapping_data" / "standardization_dict.xlsx"
    DEFAULT_REF_PATH = Path(__file__).resolve().parents[1] / "mapping_data" / "standardization_ref.xlsx"

    FALLBACK_TOKENS = {
        "编号": "no",
        "名称": "name",
        "类型": "type",
        "日期": "dt",
        "时间": "time",
        "金额": "amt",
        "余额": "bal",
        "利率": "rate",
        "币种": "currency",
        "状态": "stat",
        "笔数": "cnt",
        "差异": "diff",
        "发放": "issue",
        "到期": "mature",
        "风险": "risk",
        "等级": "lvl",
        "手机": "phone",
        "证件": "id",
    }

    def __init__(self, dict_path: Path | None = None, ref_path: Path | None = None) -> None:
        self.dict_path = dict_path or self.DEFAULT_DICT_PATH
        self.ref_path = ref_path or self.DEFAULT_REF_PATH
        self.cn_to_ref = self._load_ref_fields()
        self.token_map = self._load_token_map()

    def recommend_name(
        self,
        cn_name: str,
        explicit_name: str = "",
        source_field: str = "",
    ) -> dict:
        normalized_cn = cn_name.strip()
        explicit = explicit_name.strip().lower()
        source = source_field.strip().lower()

        if explicit:
            exact = self.cn_to_ref.get(normalized_cn)
            if exact and exact["field_name"] != explicit:
                return {
                    "name": explicit,
                    "source": "explicit_name",
                    "message": f"字段“{normalized_cn}”已填写英文名 {explicit}，与贯标参考 {exact['field_name']} 不同。",
                    "reference_name": exact["field_name"],
                    "reference_type": exact["field_type"],
                }
            return {
                "name": explicit,
                "source": "explicit_name",
                "message": "",
                "reference_name": "",
                "reference_type": "",
            }

        exact = self.cn_to_ref.get(normalized_cn)
        if exact:
            message = f"字段“{normalized_cn}”已匹配贯标字段 {exact['field_name']}（{exact['field_type']}）。"
            if source and source != exact["field_name"]:
                message += f" 来源字段为 {source}，目标字段按贯标名输出。"
            return {
                "name": exact["field_name"],
                "source": "reference",
                "message": message,
                "reference_name": exact["field_name"],
                "reference_type": exact["field_type"],
            }

        token_name = self._build_name_from_tokens(normalized_cn)
        if token_name:
            if source:
                message = f"字段“{normalized_cn}”未命中贯标参考，词典建议英文名 {token_name}，当前暂沿用来源字段 {source}。"
                return {
                    "name": source,
                    "source": "source_field",
                    "message": message,
                    "reference_name": token_name,
                    "reference_type": "",
                }
            message = f"字段“{normalized_cn}”未命中贯标参考，按词典推荐英文名 {token_name}。"
            return {
                "name": token_name,
                "source": "token_dictionary",
                "message": message,
                "reference_name": "",
                "reference_type": "",
            }

        if source:
            return {
                "name": source,
                "source": "source_field",
                "message": f"字段“{normalized_cn}”未命中贯标参考，暂沿用来源字段 {source}。",
                "reference_name": "",
                "reference_type": "",
            }

        return {
            "name": "",
            "source": "unresolved",
            "message": f"字段“{normalized_cn}”暂未生成贯标英文名。",
            "reference_name": "",
            "reference_type": "",
        }

    def _load_ref_fields(self) -> dict[str, dict]:
        if not self.ref_path.exists():
            return {}
        workbook = openpyxl.load_workbook(self.ref_path, data_only=True)
        sheet = workbook.active
        result: dict[str, dict] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cn_name = self._cell_text(row[3] if len(row) > 3 else "")
            field_name = self._normalize_field_name(row[2] if len(row) > 2 else "")
            field_type = self._cell_text(row[4] if len(row) > 4 else "")
            if cn_name and field_name and cn_name not in result:
                result[cn_name] = {
                    "field_name": field_name,
                    "field_type": field_type or "STRING",
                }
        return result

    def _load_token_map(self) -> dict[str, str]:
        token_map = dict(self.FALLBACK_TOKENS)
        if not self.dict_path.exists():
            return token_map
        workbook = openpyxl.load_workbook(self.dict_path, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cn_name = self._cell_text(row[0] if len(row) > 0 else "")
            short_name = self._normalize_field_name(row[2] if len(row) > 2 else "")
            full_name = self._normalize_field_name(row[1] if len(row) > 1 else "")
            if cn_name:
                token_map[cn_name] = short_name or full_name or token_map.get(cn_name, "")
        return token_map

    def _build_name_from_tokens(self, cn_name: str) -> str:
        remaining = cn_name
        tokens: list[str] = []
        for cn_token, en_token in sorted(self.token_map.items(), key=lambda item: len(item[0]), reverse=True):
            if not cn_token or not en_token:
                continue
            if cn_token in remaining:
                remaining = remaining.replace(cn_token, " ")
                tokens.append(en_token)
        normalized_tokens = []
        seen = set()
        for token in tokens:
            if token and token not in seen:
                normalized_tokens.append(token)
                seen.add(token)
        return "_".join(normalized_tokens)

    def _normalize_field_name(self, value) -> str:
        text = self._cell_text(value).strip("`").lower()
        return re.sub(r"[^a-z0-9_]+", "_", text).strip("_")

    def _cell_text(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()
