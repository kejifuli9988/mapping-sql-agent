from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import PurePosixPath
import re
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import openpyxl

from .mapping_loader import MappingLoader
from .standardization_advisor import StandardizationAdvisor


NS_MAIN = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_REL = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
}


@dataclass
class ExcelSheet:
    name: str
    rows: list[list[str]]


class ExcelMappingParser:
    """Parse business requirement Excel workbooks into normalized mapping JSON."""

    SHEET_ALIASES = {
        "overview": {"overview", "config", "meta", "basic", "基础信息", "配置", "概览"},
        "sources": {"sources", "source", "来源表", "源表"},
        "joins": {"joins", "join", "关联", "关联关系"},
        "filters": {"filters", "filter", "过滤", "过滤条件"},
        "target_columns": {
            "target_columns",
            "targetcolumn",
            "columns",
            "mapping",
            "字段映射",
            "目标字段",
        },
    }

    HEADER_ALIASES = {
        "task_name": {"taskname", "task_name", "任务名", "任务名称"},
        "target_table": {"targettable", "target_table", "目标表", "目标表名"},
        "target_partition": {"targetpartition", "target_partition", "目标分区", "分区"},
        "name": {"name", "字段名", "表名", "名称"},
        "alias": {"alias", "别名"},
        "type": {"type", "jointype", "关联类型"},
        "right_alias": {"rightalias", "right_alias", "右表别名", "关联表别名"},
        "condition": {"condition", "joincondition", "关联条件", "条件", "表达式"},
        "expression": {"expression", "expr", "映射表达式", "字段表达式"},
    }

    def __init__(self) -> None:
        self.loader = MappingLoader()
        self.standardization_advisor = StandardizationAdvisor()

    def parse(self, content: bytes) -> dict:
        return self.parse_with_metadata(content)["mapping"]

    def parse_with_metadata(self, content: bytes) -> dict:
        try:
            return self._parse_business_requirement(content)
        except Exception as business_error:  # noqa: BLE001
            try:
                mapping = self._parse_standard_mapping(content)
                return {
                    "mapping": mapping,
                    "format": "legacy_mapping_excel",
                    "message": "未识别为 mapping_data 业务字段需求格式，已尝试兼容解析为内部 Mapping JSON。",
                    "diagnostics": [
                        f"业务字段需求解析失败：{business_error}",
                        "建议优先使用 mapping_data 中的“字段需求 / 源表结构 / 加工逻辑”业务格式。",
                    ],
                }
            except Exception as standard_error:  # noqa: BLE001
                raise ValueError(
                    "Excel 文件无法识别为 mapping_data 业务字段需求表。请使用字段需求格式："
                    "字段中文名、来源表、来源字段、DA试算逻辑等列。"
                    f"业务需求表错误：{business_error}；兼容解析错误：{standard_error}"
                ) from standard_error

    def _parse_standard_mapping(self, content: bytes) -> dict:
        workbook = self._load_workbook(content)
        mapping = {
            **self._parse_overview(workbook),
            "sources": self._parse_table_sheet(workbook, "sources", ["name", "alias"]),
            "joins": self._parse_optional_joins(workbook),
            "filters": self._parse_optional_filters(workbook),
            "target_columns": self._parse_table_sheet(
                workbook,
                "target_columns",
                ["name", "expression"],
            ),
        }
        return self.loader.validate_mapping(mapping)

    def _parse_business_requirement(self, content: bytes) -> dict:
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
        requirement_sheet = self._find_business_requirement_sheet(workbook)
        if requirement_sheet is None:
            raise ValueError("未找到包含字段中文名/DA试算逻辑/来源字段等列的业务字段需求工作表。")

        header_row, header_map = self._find_business_header(requirement_sheet)
        rows = list(requirement_sheet.iter_rows(min_row=header_row + 1, values_only=True))
        source_structures = self._parse_business_source_structures(workbook)
        logic_text = self._parse_business_logic_text(workbook)
        source_aliases: dict[str, str] = {}
        source_fields: dict[str, set[str]] = {}
        target_columns: list[dict] = []
        filters: list[str] = []
        diagnostics: list[str] = [f"识别业务字段需求工作表：{requirement_sheet.title}。"]
        standardization_notes: list[str] = []

        def register_source(table_name: str) -> str:
            table_name = table_name.strip()
            if not table_name:
                return ""
            if table_name not in source_aliases:
                source_aliases[table_name] = f"t{len(source_aliases) + 1}"
            return source_aliases[table_name]

        for table_name, meta in source_structures.items():
            alias = register_source(table_name)
            source_fields.setdefault(table_name, set()).update(meta.get("fields", set()))
            if alias and meta.get("partition_field"):
                filters.append(f"{alias}.{meta['partition_field']} = '${{biz_date}}'")

        for index, row in enumerate(rows, start=header_row + 1):
            record = self._business_record_from_row(row, header_map)
            if not any(record.values()):
                continue

            source_table = record.get("source_table", "")
            source_field = self._clean_field_name(record.get("source_field", ""))
            logic = record.get("logic", "")

            parsed_logic = self._parse_da_logic(logic)
            if parsed_logic["source_table"]:
                source_table = parsed_logic["source_table"]
            if not source_field and parsed_logic["source_field"]:
                source_field = parsed_logic["source_field"]
            if not source_table:
                source_table = self._infer_source_table(source_field, parsed_logic["expression"], source_structures)
            if not source_table:
                diagnostics.append(f"第 {index} 行未识别来源表，已跳过：{record.get('cn_name') or source_field or logic}")
                continue

            alias = register_source(source_table)
            if source_field:
                source_fields.setdefault(source_table, set()).add(source_field)

            expression = self._qualify_expression(
                parsed_logic["expression"] or source_field,
                alias,
                source_field,
            )
            if parsed_logic["filter"]:
                filters.append(self._qualify_expression(parsed_logic["filter"], alias, source_field))
            if not expression:
                diagnostics.append(f"第 {index} 行缺少来源字段或 DA 试算逻辑，已跳过：{record.get('cn_name')}")
                continue

            target_name, standard_note = self._resolve_target_name(record, source_field)
            if standard_note:
                standardization_notes.append(standard_note)
            column = {
                "name": target_name,
                "expression": expression,
                "description": record.get("cn_name", ""),
                "business_rule": record.get("business_rule", ""),
                "source_table": source_table,
                "source_field": source_field,
            }
            target_columns.append(column)

        if not source_aliases:
            raise ValueError("业务字段需求表未解析出任何来源表。")
        if not target_columns:
            raise ValueError("业务字段需求表未解析出任何目标字段。")

        joins = self._infer_business_joins(source_aliases, source_fields, source_structures, logic_text)
        schema_text = self._build_business_schema_text(source_structures, source_fields, logic_text)
        mapping = {
            "task_name": "business_requirement_mapping",
            "target_table": "dws_business_requirement_day",
            "target_partition": "dt='${biz_date}'",
            "sources": [
                {"name": table_name, "alias": alias}
                for table_name, alias in source_aliases.items()
            ],
            "joins": joins,
            "filters": self._dedupe_keep_order(filters),
            "target_columns": target_columns,
            "parse_notes": [
                "原始业务 Excel 已通过规则解析转换为标准 Mapping JSON，未调用 AI。",
                "来源表、来源字段和 DA 试算逻辑被作为确定性生成约束。",
                "系统会结合本地贯标词典和已贯标字段推荐目标字段英文名。",
                "如目标表名或关联条件与真实口径不一致，可在 Mapping 输入框中继续修改。",
            ],
        }
        validated = self.loader.validate_mapping(mapping)
        diagnostics.append(
            f"解析出 {len(source_aliases)} 张来源表、{len(target_columns)} 个目标字段、{len(joins)} 个关联条件。"
        )
        if standardization_notes:
            diagnostics.append(f"已应用 {len(standardization_notes)} 条字段贯标建议。")
        return {
            "mapping": validated,
            "format": "business_requirement_excel",
            "message": "已识别为业务字段需求 Excel，并按规则转换为标准 Mapping JSON。",
            "diagnostics": diagnostics,
            "schema_text": schema_text,
            "schema_source": "business_requirement_excel",
            "standardization_notes": standardization_notes,
        }

    def extract_text(self, content: bytes) -> str:
        workbook = self._load_workbook(content)
        sections: list[str] = []
        for sheet in workbook.values():
            sections.append(f"[sheet] {sheet.name}")
            for row in sheet.rows:
                sections.append(" | ".join((cell or "").strip() for cell in row))
            sections.append("")
        return "\n".join(sections).strip()

    def _load_workbook(self, content: bytes) -> dict[str, ExcelSheet]:
        with ZipFile(BytesIO(content)) as zf:
            shared_strings = self._load_shared_strings(zf)
            workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
            rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

            relationship_targets: dict[str, str] = {}
            for rel in rels_xml.findall("pkg:Relationship", NS_REL):
                relationship_targets[rel.attrib["Id"]] = rel.attrib["Target"]

            sheets: dict[str, ExcelSheet] = {}
            for sheet in workbook_xml.findall("main:sheets/main:sheet", NS_MAIN):
                sheet_name = sheet.attrib["name"]
                rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
                target = relationship_targets.get(rel_id)
                if not target:
                    continue

                sheet_path = str(PurePosixPath("xl") / PurePosixPath(target))
                rows = self._read_sheet_rows(zf.read(sheet_path), shared_strings)
                sheets[sheet_name] = ExcelSheet(name=sheet_name, rows=rows)

        return sheets

    def _load_shared_strings(self, zf: ZipFile) -> list[str]:
        if "xl/sharedStrings.xml" not in zf.namelist():
            return []

        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
        values: list[str] = []
        for item in root.findall("main:si", NS_MAIN):
            parts = [node.text or "" for node in item.findall(".//main:t", NS_MAIN)]
            values.append("".join(parts))
        return values

    def _read_sheet_rows(self, xml_bytes: bytes, shared_strings: list[str]) -> list[list[str]]:
        root = ET.fromstring(xml_bytes)
        rows: list[list[str]] = []
        for row in root.findall("main:sheetData/main:row", NS_MAIN):
            current: list[str] = []
            expected_index = 0
            for cell in row.findall("main:c", NS_MAIN):
                cell_ref = cell.attrib.get("r", "")
                column_index = self._column_index_from_ref(cell_ref)
                while expected_index < column_index:
                    current.append("")
                    expected_index += 1

                current.append(self._cell_value(cell, shared_strings))
                expected_index += 1

            while current and current[-1] == "":
                current.pop()

            if any(value != "" for value in current):
                rows.append(current)
        return rows

    def _cell_value(self, cell: ET.Element, shared_strings: list[str]) -> str:
        cell_type = cell.attrib.get("t")
        value_node = cell.find("main:v", NS_MAIN)
        inline_text = cell.find("main:is/main:t", NS_MAIN)

        if cell_type == "inlineStr" and inline_text is not None:
            return inline_text.text or ""
        if value_node is None:
            return ""
        if cell_type == "s":
            index = int(value_node.text or "0")
            return shared_strings[index] if 0 <= index < len(shared_strings) else ""
        return value_node.text or ""

    def _column_index_from_ref(self, cell_ref: str) -> int:
        letters = "".join(ch for ch in cell_ref if ch.isalpha())
        index = 0
        for char in letters:
            index = index * 26 + (ord(char.upper()) - ord("A") + 1)
        return max(index - 1, 0)

    def _find_sheet(self, workbook: dict[str, ExcelSheet], sheet_type: str) -> ExcelSheet | None:
        aliases = self.SHEET_ALIASES[sheet_type]
        for name, sheet in workbook.items():
            if self._normalize_key(name) in {self._normalize_key(alias) for alias in aliases}:
                return sheet
        return None

    def _parse_overview(self, workbook: dict[str, ExcelSheet]) -> dict:
        sheet = self._find_sheet(workbook, "overview")
        if sheet is None:
            raise ValueError(
                "Excel Mapping 缺少 overview/config/基础信息 工作表，无法读取任务名和目标表信息。"
            )

        values: dict[str, str] = {}
        for row in sheet.rows:
            if len(row) < 2:
                continue
            key = self._map_header(row[0])
            if key in {"task_name", "target_table", "target_partition"}:
                values[key] = row[1].strip()

        missing = [key for key in ["task_name", "target_table", "target_partition"] if key not in values]
        if missing:
            raise ValueError(f"Excel Mapping 概览页缺少关键字段：{', '.join(missing)}。")
        return values

    def _parse_table_sheet(
        self,
        workbook: dict[str, ExcelSheet],
        sheet_type: str,
        required_headers: list[str],
    ) -> list[dict]:
        sheet = self._find_sheet(workbook, sheet_type)
        if sheet is None:
            aliases = "/".join(sorted(self.SHEET_ALIASES[sheet_type]))
            raise ValueError(f"Excel Mapping 缺少 {aliases} 工作表。")

        if not sheet.rows:
            raise ValueError(f"Excel Mapping 工作表 {sheet.name} 为空。")

        header_map = self._build_header_map(sheet.rows[0])
        missing = [header for header in required_headers if header not in header_map]
        if missing:
            raise ValueError(
                f"Excel Mapping 工作表 {sheet.name} 缺少列：{', '.join(missing)}。"
            )

        records: list[dict] = []
        for row in sheet.rows[1:]:
            if not any(cell.strip() for cell in row):
                continue

            record = {}
            for canonical, index in header_map.items():
                record[canonical] = row[index].strip() if index < len(row) else ""

            if any(record.get(header, "") for header in required_headers):
                records.append({header: record.get(header, "") for header in header_map})

        if not records:
            raise ValueError(f"Excel Mapping 工作表 {sheet.name} 没有有效数据行。")
        return records

    def _parse_optional_joins(self, workbook: dict[str, ExcelSheet]) -> list[dict]:
        sheet = self._find_sheet(workbook, "joins")
        if sheet is None or not sheet.rows:
            return []

        header_map = self._build_header_map(sheet.rows[0])
        if "condition" not in header_map:
            return []

        records: list[dict] = []
        for row in sheet.rows[1:]:
            if not any(cell.strip() for cell in row):
                continue
            condition = row[header_map["condition"]].strip() if header_map["condition"] < len(row) else ""
            if not condition:
                continue
            record = {
                "type": self._safe_value(row, header_map.get("type"), default="left"),
                "right_alias": self._safe_value(row, header_map.get("right_alias")),
                "condition": condition,
            }
            if record["right_alias"]:
                records.append(record)
        return records

    def _parse_optional_filters(self, workbook: dict[str, ExcelSheet]) -> list[str]:
        sheet = self._find_sheet(workbook, "filters")
        if sheet is None or not sheet.rows:
            return []

        header_map = self._build_header_map(sheet.rows[0])
        candidate_index = header_map.get("condition")
        if candidate_index is None:
            candidate_index = 0

        values: list[str] = []
        for row in sheet.rows[1:]:
            if candidate_index < len(row):
                item = row[candidate_index].strip()
                if item:
                    values.append(item)
        return values

    def _build_header_map(self, header_row: list[str]) -> dict[str, int]:
        header_map: dict[str, int] = {}
        for index, raw_header in enumerate(header_row):
            canonical = self._map_header(raw_header)
            if canonical:
                header_map[canonical] = index
        return header_map

    def _map_header(self, raw_header: str) -> str | None:
        normalized = self._normalize_key(raw_header)
        for canonical, aliases in self.HEADER_ALIASES.items():
            if normalized in {self._normalize_key(alias) for alias in aliases}:
                return canonical
        return normalized or None

    def _normalize_key(self, value: str) -> str:
        return re.sub(r"[\s_\-:：]+", "", value.strip().lower())

    def _safe_value(self, row: list[str], index: int | None, default: str = "") -> str:
        if index is None or index >= len(row):
            return default
        return row[index].strip() or default

    def _find_business_requirement_sheet(self, workbook: openpyxl.Workbook):
        for sheet in workbook.worksheets:
            try:
                _, header_map = self._find_business_header(sheet)
            except ValueError:
                continue
            if "cn_name" in header_map and ("logic" in header_map or "source_field" in header_map):
                return sheet
        return None

    def _find_business_header(self, sheet) -> tuple[int, dict[str, int]]:
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [self._cell_to_text(value) for value in row]
            header_map: dict[str, int] = {}
            for index, value in enumerate(values):
                canonical = self._map_business_header(value)
                if canonical:
                    header_map[canonical] = index
            if "cn_name" in header_map and ("logic" in header_map or "source_field" in header_map):
                return row_index, header_map
        raise ValueError(f"工作表 {sheet.title} 未找到业务字段需求表头。")

    def _map_business_header(self, raw_header: str) -> str | None:
        normalized = self._normalize_key(raw_header)
        aliases = {
            "field_no": {"序号", "编号", "no"},
            "en_name": {"字段英文名", "英文字段名", "目标字段英文名", "字段名"},
            "actual_name": {"实际字段名", "实现字段名"},
            "actual_type": {"实际类型", "字段类型", "类型"},
            "cn_name": {"字段中文名", "中文字段名", "字段中文名称", "字段名称"},
            "business_rule": {"业务口径说明", "业务口径", "口径说明"},
            "source_system": {"来源系统", "源系统"},
            "source_table": {"来源表", "源表", "表名"},
            "source_field": {"来源字段", "源字段"},
            "logic": {"DA试算逻辑", "试算逻辑", "加工逻辑", "字段表达式"},
            "remark": {"备注", "说明"},
            "diff": {"实现与需求差异", "差异"},
        }
        for canonical, values in aliases.items():
            if normalized in {self._normalize_key(item) for item in values}:
                return canonical
        return None

    def _business_record_from_row(self, row: tuple, header_map: dict[str, int]) -> dict[str, str]:
        result = {}
        for key, index in header_map.items():
            result[key] = self._cell_to_text(row[index] if index < len(row) else "")
        return result

    def _parse_business_source_structures(self, workbook: openpyxl.Workbook) -> dict[str, dict]:
        structures: dict[str, dict] = {}
        for sheet in workbook.worksheets:
            headers = None
            header_row = 0
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [self._cell_to_text(value) for value in row]
                mapped = {self._normalize_key(value): idx for idx, value in enumerate(values) if value}
                if "表名" in mapped and ("关键字段" in mapped or "分区字段" in mapped):
                    headers = mapped
                    header_row = row_index
                    break
            if headers is None:
                continue
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                values = [self._cell_to_text(value) for value in row]
                table_name = values[headers["表名"]].strip() if headers["表名"] < len(values) else ""
                if not table_name:
                    continue
                key_fields = values[headers.get("关键字段", -1)] if headers.get("关键字段", -1) < len(values) else ""
                partition_field = values[headers.get("分区字段", -1)] if headers.get("分区字段", -1) < len(values) else ""
                description = values[headers.get("中文说明", -1)] if headers.get("中文说明", -1) < len(values) else ""
                frequency = values[headers.get("更新频率", -1)] if headers.get("更新频率", -1) < len(values) else ""
                fields = {
                    self._clean_field_name(item)
                    for item in re.split(r"[,，、\s]+", key_fields)
                    if self._clean_field_name(item)
                }
                structures[table_name] = {
                    "fields": fields,
                    "partition_field": self._clean_field_name(partition_field),
                    "description": description,
                    "frequency": frequency,
                }
        return structures

    def _parse_business_logic_text(self, workbook: openpyxl.Workbook) -> str:
        chunks: list[str] = []
        for sheet in workbook.worksheets:
            if "逻辑" not in sheet.title:
                continue
            for row in sheet.iter_rows(values_only=True):
                text = " ".join(self._cell_to_text(value) for value in row if self._cell_to_text(value))
                if text:
                    chunks.append(text)
        return "\n".join(chunks)

    def _build_business_schema_text(
        self,
        source_structures: dict[str, dict],
        source_fields: dict[str, set[str]],
        logic_text: str,
    ) -> str:
        lines: list[str] = [
            "业务 Excel 自动提取的生成前表结构上下文",
            "来源：字段需求 / 源表结构 / 加工逻辑",
            "",
        ]
        table_names = self._dedupe_keep_order([*source_structures.keys(), *source_fields.keys()])
        for table_name in table_names:
            meta = source_structures.get(table_name, {})
            fields = self._dedupe_keep_order(
                [
                    *sorted(meta.get("fields", set())),
                    *sorted(source_fields.get(table_name, set())),
                    meta.get("partition_field", ""),
                ]
            )
            lines.append(f"[table] {table_name}")
            if meta.get("description"):
                lines.append(f"表用途：{meta['description']}")
            if meta.get("frequency"):
                lines.append(f"更新频率：{meta['frequency']}")
            if meta.get("partition_field"):
                lines.append(f"分区字段：{meta['partition_field']}")
            lines.append("字段名 | 类型 | 说明")
            for field_name in fields:
                if not field_name:
                    continue
                description = "分区字段" if field_name == meta.get("partition_field") else self._infer_field_description(field_name)
                lines.append(f"{field_name} | STRING | {description}")
            lines.append("")
        if logic_text:
            lines.append("[加工逻辑]")
            lines.append(logic_text)
        return "\n".join(lines).strip()

    def _infer_field_description(self, field_name: str) -> str:
        lowered = field_name.lower()
        if lowered in {"dt", "biz_date", "data_dt"} or lowered.endswith("_dt"):
            return "日期字段"
        if lowered.endswith("_time"):
            return "时间字段"
        if any(token in lowered for token in ("amt", "bal", "amount", "balance")):
            return "金额或余额指标字段"
        if any(token in lowered for token in ("rate", "ratio")):
            return "比例或利率指标字段"
        if any(token in lowered for token in ("id", "no", "code")):
            return "标识或关联字段"
        if lowered.endswith("_name") or "name" in lowered:
            return "名称描述字段"
        if any(token in lowered for token in ("type", "status", "flag")):
            return "分类状态字段"
        return "来源字段"

    def _parse_da_logic(self, logic: str) -> dict[str, str]:
        text = logic.strip()
        result = {"expression": text, "source_table": "", "source_field": "", "filter": ""}
        if not text:
            return result

        select_match = re.search(
            r"select\s+(.+?)\s+from\s+([`\w.]+)",
            text,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if select_match:
            expression = select_match.group(1).strip()
            result["expression"] = expression
            result["source_field"] = self._first_field_token(expression)
            result["source_table"] = select_match.group(2).strip("` ")
            return result

        where_match = re.search(r"\bwhere\b", text, flags=re.IGNORECASE)
        if where_match:
            result["expression"] = text[: where_match.start()].strip()
            result["filter"] = text[where_match.end() :].strip()
        result["source_field"] = self._first_field_token(result["expression"])
        return result

    def _first_field_token(self, expression: str) -> str:
        for token in re.findall(r"[A-Za-z_][A-Za-z0-9_]*", expression):
            if token.upper() in {"SUM", "COUNT", "MAX", "MIN", "AVG", "COALESCE", "CASE", "WHEN", "THEN", "ELSE", "END", "SELECT", "FROM", "WHERE"}:
                continue
            return self._clean_field_name(token)
        return ""

    def _infer_source_table(self, source_field: str, expression: str, structures: dict[str, dict]) -> str:
        candidates = {source_field}
        candidates.update(self._clean_field_name(token) for token in re.findall(r"[A-Za-z_][A-Za-z0-9_]*", expression))
        candidates.discard("")
        for table_name, meta in structures.items():
            if candidates & meta.get("fields", set()):
                return table_name
        return ""

    def _qualify_expression(self, expression: str, alias: str, source_field: str) -> str:
        text = expression.strip()
        if not text:
            return ""
        if re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", text):
            return f"{alias}.{text}"

        protected = {
            "SUM", "COUNT", "MAX", "MIN", "AVG", "COALESCE", "NULLIF",
            "CASE", "WHEN", "THEN", "ELSE", "END", "DISTINCT",
            "SELECT", "FROM", "WHERE", "IN", "AND", "OR", "AS", "IS", "NOT", "NULL",
        }
        source_candidates = {source_field} if source_field else set()
        source_candidates.update(
            token for token in re.findall(r"[A-Za-z_][A-Za-z0-9_]*", text)
            if token.upper() not in protected
        )
        parts = re.split(r"('(?:''|[^'])*')", text)
        for part_index in range(0, len(parts), 2):
            segment = parts[part_index]
            for field in sorted(source_candidates, key=len, reverse=True):
                if not field:
                    continue
                segment = re.sub(
                    rf"(?<![.\w]){re.escape(field)}(?![\w])",
                    f"{alias}.{field}",
                    segment,
                )
            parts[part_index] = segment
        return "".join(parts)

    def _resolve_target_name(self, record: dict[str, str], source_field: str) -> tuple[str, str]:
        explicit_name = ""
        for key in ("en_name", "actual_name"):
            value = self._clean_field_name(record.get(key, ""))
            if value:
                explicit_name = value.lower()
                break
        recommendation = self.standardization_advisor.recommend_name(
            cn_name=record.get("cn_name", ""),
            explicit_name=explicit_name,
            source_field=source_field,
        )
        if recommendation["name"]:
            return recommendation["name"], recommendation["message"]
        if source_field:
            return source_field.lower(), recommendation["message"]
        cn_name = record.get("cn_name", "")
        return f"field_{abs(hash(cn_name)) % 10000}", recommendation["message"]

    def _infer_business_joins(
        self,
        source_aliases: dict[str, str],
        source_fields: dict[str, set[str]],
        source_structures: dict[str, dict],
        logic_text: str,
    ) -> list[dict]:
        tables = list(source_aliases)
        if len(tables) < 2:
            return []

        joins = []
        base_table = tables[0]
        base_alias = source_aliases[base_table]
        base_fields = source_fields.get(base_table, set()) | source_structures.get(base_table, {}).get("fields", set())
        for table_name in tables[1:]:
            alias = source_aliases[table_name]
            fields = source_fields.get(table_name, set()) | source_structures.get(table_name, {}).get("fields", set())
            condition = self._infer_join_condition(base_alias, base_fields, alias, fields, logic_text, table_name)
            joins.append({"type": "left", "right_alias": alias, "condition": condition})
        return joins

    def _infer_join_condition(self, base_alias: str, base_fields: set[str], alias: str, fields: set[str], logic_text: str, table_name: str) -> str:
        common = sorted((base_fields & fields) - {""})
        if common:
            field = common[0]
            return f"{base_alias}.{field} = {alias}.{field}"
        if "ecif_cust_id" in base_fields and ("cust_id" in fields or "cont" in table_name.lower() or "loan" in table_name.lower()):
            return f"{base_alias}.ecif_cust_id = {alias}.cust_id"
        if "cust_id" in base_fields and "cust_id" in fields:
            return f"{base_alias}.cust_id = {alias}.cust_id"
        if "客户编号" in logic_text and "cust_id" in logic_text:
            left_field = "ecif_cust_id" if "ecif_cust_id" in base_fields else "cust_id"
            return f"{base_alias}.{left_field} = {alias}.cust_id"
        return f"{base_alias}.id = {alias}.id"

    def _dedupe_keep_order(self, values: list[str]) -> list[str]:
        result: list[str] = []
        seen = set()
        for value in values:
            normalized = value.strip()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            result.append(normalized)
        return result

    def _cell_to_text(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _clean_field_name(self, value: str) -> str:
        text = value.strip().strip("`")
        text = re.sub(r"[^A-Za-z0-9_]+", "", text)
        return text
